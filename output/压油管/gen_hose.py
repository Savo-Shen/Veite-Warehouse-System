# -*- coding: utf-8 -*-
"""生成「压油管」基础资料底稿 xlsx（通径口径按用户实际编号）"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sizes import (BORE, BY_CODE, CODES, label, LAYERS, METRIC_L, METRIC_S, BSP,
                   JIC, NPT, ORFS, JIS30, FLANGE61, FLANGE62, HOSE_COUNT, HOSE_LOC,
                   FERRULE_GROUPS, price_of, PRICE_KNOWN)
import fittings as FIT

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "压油管_接头油管底稿.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
FILL_FILL = PatternFill("solid", fgColor="FFF2CC")   # 待填列
REF_FILL = PatternFill("solid", fgColor="EAF1F8")    # 参考列
DONE_FILL = PatternFill("solid", fgColor="E2EFDA")   # 已盘数据
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_sheet(ws, headers, rows, fill_cols=(), ref_cols=(), done_cols=(), widths=None):
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    for r in rows:
        ws.append(r)
    for r in range(2, ws.max_row + 1):
        for i in range(1, len(headers) + 1):
            c = ws.cell(row=r, column=i)
            c.border, c.font = BORDER, Font(size=10)
            c.alignment = Alignment(vertical="center")
            if headers[i - 1] in fill_cols:
                c.fill = FILL_FILL
            elif headers[i - 1] in done_cols:
                c.fill = DONE_FILL
            elif headers[i - 1] in ref_cols:
                c.fill = REF_FILL
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


LAYER_BY_CODE = {l[0]: l for l in LAYERS}
OD_IDX = {"01": 4, "02": 5, "04": 6}
WP_IDX = {"01": 7, "02": 8, "04": 9}


# ============================================================
# 01-油管（已盘）
# ============================================================
def hose_rows():
    rows = []
    for code in sorted(HOSE_COUNT, key=lambda c: (CODES.index(c[:2]), c[2:])):
        bcode, lcode = code[:2], code[2:]
        b = BY_CODE[bcode]
        lay = LAYER_BY_CODE[lcode]
        pieces = HOSE_COUNT[code]
        od = b[OD_IDX[lcode]] if lcode in OD_IDX else None
        wp = b[WP_IDX[lcode]] if lcode in WP_IDX else None
        price, src = price_of(bcode, lcode)
        rows.append([
            code, f'高压油管 {code} {label(bcode)} {lay[1]}',
            lay[1], lay[2], lay[3],
            b[1], b[2], bcode, b[3], od, wp, round(wp * 4, 1) if wp else None, b[10],
            "米", sum(pieces), len(pieces),
            " + ".join(f"{p}m" for p in pieces),
            HOSE_LOC,
            "",
            price, src, round(price * sum(pieces), 2) if price else None,
            "余料不可合并使用" if len(pieces) > 1 else "",
        ])
    return rows


HOSE_HEADERS = [
    "SKU编码", "SKU名称", "层数", "结构名称", "对应标准",
    "通径(英寸)", "俗称", "内径码", "内径mm", "外径mm(参考)",
    "工作压力MPa(参考)", "爆破压力MPa(参考)", "最小弯曲半径mm(参考)",
    "单位", "库存合计(米)", "段数", "分段长度", "库位",
    "★品牌", "成本价(元/米)", "价格来源", "库存金额(元)", "备注",
]
HOSE_DONE = {"库存合计(米)", "段数", "分段长度", "库位", "成本价(元/米)",
             "价格来源", "库存金额(元)"}
HOSE_FILL = {"★品牌"}
HOSE_REF = {"内径mm", "外径mm(参考)", "工作压力MPa(参考)", "爆破压力MPa(参考)",
            "最小弯曲半径mm(参考)"}

# ============================================================
# 02-接头
# ============================================================
FITTING_SERIES = [
    ("24L-F", "公制24°锥内螺纹 轻系列", "内螺纹(母)", "24°球面/内锥密封", "公制 M(DIN 2353-L)", METRIC_L, "最常用"),
    ("24S-F", "公制24°锥内螺纹 重系列", "内螺纹(母)", "24°球面/内锥密封", "公制 M(DIN 2353-S)", METRIC_S, "常用"),
    ("24L-M", "公制24°锥外螺纹 轻系列", "外螺纹(公)", "24°锥+O型圈", "公制 M(DIN 2353-L)", METRIC_L, "常用"),
    ("24S-M", "公制24°锥外螺纹 重系列", "外螺纹(公)", "24°锥+O型圈", "公制 M(DIN 2353-S)", METRIC_S, "常用"),
    ("BSP-F", "英制60°锥内螺纹", "内螺纹(母)", "60°内锥密封", "英制 G(BSPP)", BSP, "最常用"),
    ("BSP-M", "英制60°锥外螺纹", "外螺纹(公)", "60°外锥密封", "英制 G(BSPP)", BSP, "常用"),
    ("JIC-F", "美制74°锥内螺纹(JIC)", "内螺纹(母)", "74°内锥密封", "美制 UNF", JIC, "常用"),
    ("JIC-M", "美制74°锥外螺纹(JIC)", "外螺纹(公)", "74°外锥密封", "美制 UNF", JIC, "常用"),
    ("NPT-M", "美制锥管螺纹外螺纹(NPT)", "外螺纹(公)", "螺纹锥密封", "美制 NPT", NPT, "常用"),
    ("ORFS-F", "ORFS内螺纹(O型圈平面)", "内螺纹(母)", "O型圈平面密封", "美制 UNF", ORFS, "备选"),
    ("ORFS-M", "ORFS外螺纹(O型圈平面)", "外螺纹(公)", "O型圈平面密封", "美制 UNF", ORFS, "备选"),
    ("JIS-F", "日标30°锥内螺纹(PF)", "内螺纹(母)", "30°内锥密封", "日标 PF", JIS30, "备选"),
    ("FL61", "SAE法兰 3000PSI(61系列)", "法兰", "O型圈端面密封", "SAE J518 code61", FLANGE61, "备选"),
    ("FL62", "SAE法兰 6000PSI(62系列)", "法兰", "O型圈端面密封", "SAE J518 code62", FLANGE62, "备选"),
]
ANGLES = [("直", "直头", "最常用"), ("45", "45°弯头", "常用"), ("90", "90°弯头", "最常用")]
RANK = {"最常用": 2, "常用": 1, "备选": 0}


# 02-接头的行生成已搬到 fittings.py（改用现场口径：螺纹 + A/C/D型 + 弯 + 芯/面）

# ============================================================
# 03-扣压外套
# ============================================================
SKIN = [("非剥皮", "非剥皮式(整皮扣压)", "最常用"), ("剥皮", "剥皮式(需剥外胶)", "常用")]


def ferrule_rows():
    rows = []
    for fcode, fname, layer, codes, split, fhot in FERRULE_GROUPS:
        skins = SKIN if split else [("", "", "最常用")]
        for bcode in codes:
            b = BY_CODE[bcode]
            for skcode, skname, skhot in skins:
                hot = {2: "最常用", 1: "常用", 0: "备选"}[min(RANK[fhot], RANK[skhot])]
                sku = f"{fcode}-{bcode}" + (f"-{skcode}" if skcode else "")
                rows.append([
                    sku,
                    f'扣压外套 {fname}{" " + skname if skname else ""} {label(bcode)}',
                    fcode, fname, layer, skname or "不分", b[1], b[2], bcode, "个", hot,
                    "", "", "", "", "", "",
                ])
    return rows


FER_HEADERS = [
    "建议SKU编码", "建议SKU名称", "外套代号", "外套名称", "适用层数", "剥皮方式",
    "配管通径(英寸)", "俗称", "内径码", "单位", "常用度",
    "★厂家代号", "★品牌", "★现有库存(个)", "★库位", "★成本价(元/个)", "★备注",
]
FER_FILL = {"★厂家代号", "★品牌", "★现有库存(个)", "★库位", "★成本价(元/个)", "★备注"}


def crimp_rows():
    rows = []
    for _fc, _fn, layer, codes, _sp, _ho in FERRULE_GROUPS:
        for bcode in codes:
            b = BY_CODE[bcode]
            rows.append([layer, b[1], b[2], bcode, b[3], "", "", "", "", "", ""])
    return rows


CRIMP_HEADERS = [
    "层数", "通径(英寸)", "俗称", "内径码", "内径mm",
    "★扣压直径mm", "★模具号", "★剥胶长度mm", "★插入深度mm", "★压机档位", "★备注",
]
CRIMP_FILL = {"★扣压直径mm", "★模具号", "★剥胶长度mm", "★插入深度mm", "★压机档位", "★备注"}

# ============================================================
wb = Workbook()

ws = wb.active
ws.title = "00-说明"
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 118
total_m = sum(sum(v) for v in HOSE_COUNT.values())
total_p = sum(len(v) for v in HOSE_COUNT.values())
lines = [
    ("H", "压油管模块 · 接头与油管基础资料"),
    ("", ""),
    ("H2", "一、编号规则（按你现场实际用的）"),
    ("", "胶管编号 4 位：前两位 = 内径mm，后两位 = 层数。例：1302 = 内径13(1/2\" 四分) 二层钢丝。"),
    ("", "所以通径一律用内径码 06 / 08 / 10 / 13 / 16 / 19 / 22 / 25 / 32 / 38 / 51 表示，不再写 DN12、DN31.5。"),
    ("", "扣压外套的编码跟着用同一套内径码，比如 F12-13-非剥皮 = 一二层管用 配四分管 非剥皮式。"),
    ("", ""),
    ("H2", "一之二、接头编号规则（2026-08-23 按你现场标签改的）"),
    ("", "接头不跟内径码走——现场标签上本来就只写螺纹，不写配几分管。编码结构跟手写完全一致："),
    ("", "    螺纹 + 型 + [弯] + 芯/面"),
    ("", "    22×1.5 A型芯      →  M22x1.5-A-芯"),
    ("", "    18×1.5 C型弯面    →  M18x1.5-C-弯-面"),
    ("", "    3分英制A型面      →  G3/8-A-面"),
    ("", "    3分美制D型芯      →  JIC3/8-D-芯"),
    ("", "· 型 = A / C / D，指接头里面的密封座形状，不是弯头角度。标准叫法见「02c-型对照」。"),
    ("", "· 芯 = 公头(外螺纹外锥)；面 = 母头(内螺纹带活动螺母)。"),
    ("", "· 不写「弯」就是直头。"),
    ("", "· 「可配管通径」只做参考列，不进编码——公制螺纹有轻/重系列两解，"),
    ("", "  比如 M18×1.5 既是轻系列配三分管、也是重系列配二分半管，光看标签分不出来。"),
    ("", ""),
    ("H2", "二、进度"),
    ("", f"胶管：已盘完（2026-08-23），{len(HOSE_COUNT)} 个规格、{total_p} 段、合计 {total_m} 米，全部在库位 {HOSE_LOC}。见「01-油管(已盘)」。"),
    ("", "接头、扣压外套：待盘，见「02-接头」「03-扣压外套」，带 ★ 的黄色列是要填的。"),
    ("", ""),
    ("H2", "三、分段长度要单独记"),
    ("", "余料不能合并使用——1602 有 10+10+4+8 米四段，接一张要 12 米的单子，虽然合计 32 米但一段都不够。"),
    ("", "所以「分段长度」列按段记，不要只留合计。系统里这个也要落到单独的字段，不能只存一个数量。"),
    ("", ""),
    ("H2", "四、页签说明"),
    ("", "01-油管(已盘)  绿色列是已经盘出来的实际数据，黄色列还差品牌和成本价。"),
    ("", "02-接头        按「螺纹 × A/C/D型 × 直/弯 × 芯/面」展开，单位个。系统导入用这页。"),
    ("", "02b-接头盘点矩阵 同样的数据摊成表格：一行一个螺纹，一列一个型/角度/芯面组合，手工点数用这页。"),
    ("", "02c-型对照ACD   A/C/D 各是什么密封座、对应哪个标准、怎么在现场认出来。"),
    ("", "03-扣压外套    皮子，按「适用层数 × 通径 × 剥皮方式」展开，单位个。"),
    ("", "04-通径对照表  内径码/英寸/俗称/外径/工作压力/弯曲半径换算，接电话时直接查。"),
    ("", "05-螺纹对照表  各接头系列在各通径下的常配螺纹。"),
    ("", "06-现有库位    从生产库(192.168.1.4)导出的现有库位。"),
    ("", "07-扣压参数    扣压直径/模具号/剥胶长度，机器和厂牌相关，现场实测后填。"),
    ("", "08-油管价格    成本价 + 三档售价参考。"),
    ("", ""),
    ("H2", "五、价格怎么来的"),
    ("", f"你 2026-08-23 报的 {len(PRICE_KNOWN)} 个实际进价标「实价」，其余标「推算」——按实价最小二乘拟合出来的："),
    ("", "  1层 = 1.634 + 0.5798×内径mm（4 个实价点，最大偏差 4.4%）"),
    ("", "  4层 = -1.068 + 1.3281×内径mm（3 个实价点，最大偏差 5.4%）"),
    ("", "  2层 在你给的区间里正好比 1层 贵 1.0 元/米，外推到大口径时按口径放大（粗管第二层钢丝用料更多）。"),
    ("", "  3层 没有实价，取 2层 与 4层 的中值。"),
    ("", "推算值取整到 0.5 元，就是提示这是估价。**误差按 ±15% 看**，真要报价还是问一下供货商。"),
    ("", "「售价×1.5 / ×1.8 / ×2.0」是 Excel 公式，改成本价会自动跟着变；实际加价率你自己定。"),
    ("", ""),
    ("H2", "六、待确认"),
    ("", "· 2503 的「03」是三层钢丝缠绕吗？三层不是欧标序列里的常规品，确认一下你们是不是这么叫。"),
    ("", "· 22（7/8\"）不在 EN 853 / EN 856 序列里，国内厂家有做但参数各家不同，"),
    ("", "  所以它的外径、工作压力我留空了，接头螺纹除英制 G7/8 外也没敢填，按实物补。"),
    ("", "· 表里的外径、压力、弯曲半径是欧标参考值，对外报参数以厂家样本为准。"),
    ("", "· A/C/D 对应的标准密封形式（60°锥 / 24°锥 / 带O圈）是照片目测的，还没实测，见「02c-型对照」。"),
    ("", "· 「纸上见过」这列是照着你手写那张纸打的勾，字迹有几处我读不准，请核一遍再往下建档。"),
    ("", "· 「弯」目前只有一档。如果你们其实分 45°弯和 90°弯，说一声，我把这一维拆成两个值。"),
    ("", "· 手写纸上还有「长六角」「短」「φ10双层」「铰细锥变头」几类没建档："),
    ("", "  前两个是本体长度变体，要在接头表上再加一列；变头是转换接头不扣胶管，得单开一个品类。"),
]
r = 1
for kind, text in lines:
    c = ws.cell(row=r, column=2, value=text)
    if kind == "H":
        c.font = Font(size=16, bold=True, color="1F4E79"); ws.row_dimensions[r].height = 26
    elif kind == "H2":
        c.font = Font(size=12, bold=True, color="1F4E79"); ws.row_dimensions[r].height = 22
    else:
        c.font = Font(size=10)
    c.alignment = Alignment(vertical="center")
    r += 1

ws = wb.create_sheet("01-油管(已盘)")
write_sheet(ws, HOSE_HEADERS, hose_rows(), HOSE_FILL, HOSE_REF, HOSE_DONE,
            widths=[10, 30, 8, 22, 34, 11, 8, 8, 9, 13, 16, 16, 18, 6, 14, 7, 22, 10, 12, 14, 10, 14, 20])

ws = wb.create_sheet("02-接头")
write_sheet(ws, FIT.FIT_HEADERS, FIT.fitting_rows(), FIT.FIT_FILL, FIT.FIT_REF, set(),
            widths=FIT.FIT_WIDTHS)

ws = wb.create_sheet("02b-接头盘点矩阵")
write_sheet(ws, FIT.matrix_headers(), FIT.matrix_rows(),
            set(FIT.matrix_headers()[4:]), {"可配管通径(参考)"}, set(),
            widths=FIT.MATRIX_WIDTHS)

ws = wb.create_sheet("02c-型对照ACD")
write_sheet(ws, FIT.SEAT_HEADERS, FIT.seat_rows(), set(),
            {"密封形式(标准)", "对应标准/代号"}, set(),
            widths=[6, 38, 26, 32, 12, 60])

ws = wb.create_sheet("03-扣压外套")
write_sheet(ws, FER_HEADERS, ferrule_rows(), FER_FILL, set(), set(),
            widths=[18, 46, 10, 26, 11, 18, 12, 8, 8, 6, 9, 13, 12, 14, 12, 14, 18])

ws = wb.create_sheet("04-通径对照表")
bh = ["内径码", "通径(英寸)", "俗称", "内径mm", "1层外径mm", "2层外径mm", "4层外径mm",
      "1层工作压力MPa", "2层工作压力MPa", "4层工作压力MPa", "最小弯曲半径mm"]
write_sheet(ws, bh, [list(b) for b in BORE], set(), set(bh[3:]), set(),
            widths=[8, 11, 8, 9, 12, 12, 12, 16, 16, 16, 16])

ws = wb.create_sheet("05-螺纹对照表")
th = ["接头系列", "系列代号", "公/母", "密封形式", "螺纹标准"] + \
     [f'{b[0]} {b[2]} {b[1]}"' for b in BORE]
write_sheet(ws, th, [[sn, sc, g, sl, ts] + [tm.get(b[0], "") for b in BORE]
                     for sc, sn, g, sl, ts, tm, _ in FITTING_SERIES], set(), set(), set(),
            widths=[26, 10, 12, 18, 20] + [14] * len(BORE))

ws = wb.create_sheet("06-现有库位")
loc_rows = []
lf = os.path.join(HERE, "locations.tsv")
if os.path.exists(lf):
    with open(lf, encoding="utf-8") as f:
        for line in f:
            p = line.rstrip("\n").split("\t")
            if len(p) >= 2 and p[0] != "location_code":
                loc_rows.append([p[0], p[1], "← 胶管全部在这" if p[0] == HOSE_LOC else ""])
write_sheet(ws, ["库位码", "库位名称", "说明"], loc_rows, set(), set(), set(),
            widths=[14, 34, 40])

ws = wb.create_sheet("07-扣压参数")
write_sheet(ws, CRIMP_HEADERS, crimp_rows(), CRIMP_FILL, set(), set(),
            widths=[10, 11, 8, 8, 9, 15, 12, 15, 15, 13, 24])

ws = wb.create_sheet("08-油管价格")
LCODES = [("01", "1层"), ("02", "2层"), ("03", "3层"), ("04", "4层")]
price_rows = []
for bcode in CODES:
    b = BY_CODE[bcode]
    for lcode, lname in LCODES:
        pr, src = price_of(bcode, lcode)
        if pr is None:
            continue
        code = bcode + lcode
        stock = sum(HOSE_COUNT[code]) if code in HOSE_COUNT else None
        price_rows.append([code, b[1], b[2], bcode, lname, pr, src, stock,
                           round(pr * stock, 2) if stock else None])
PRICE_HEADERS = ["代号", "通径(英寸)", "俗称", "内径码", "层数",
                 "成本价(元/米)", "来源", "现有库存(米)", "库存金额(元)",
                 "售价×1.5", "售价×1.8", "售价×2.0"]
write_sheet(ws, PRICE_HEADERS, [r + [None, None, None] for r in price_rows],
            set(), set(), {"成本价(元/米)", "来源", "库存金额(元)"},
            widths=[8, 11, 8, 8, 7, 14, 8, 14, 14, 11, 11, 11])
for i in range(len(price_rows)):
    r = i + 2
    for col, mul in (("J", 1.5), ("K", 1.8), ("L", 2.0)):
        ws[f"{col}{r}"] = f"=ROUND(F{r}*{mul},1)"

wb.save(OUT)
print("saved:", OUT)
for s in wb.sheetnames:
    print(f"  {s}: {wb[s].max_row - 1} 行")
print("合计", total_m, "米 /", total_p, "段")
